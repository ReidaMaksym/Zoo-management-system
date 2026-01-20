from openpyxl import load_workbook, Workbook
from src.zoo_manager import ZooManager
from src.user import User
from src.section import ZooSection


def get_data_from_sheet(sheet_name: str, work_book: Workbook) -> list:

    if sheet_name not in work_book.sheetnames:
        print(f"The sheet '{sheet_name}' doesn't exist")
        return []
    
    work_sheet = work_book[sheet_name]

    entities = []

    for i in range(2, work_sheet.max_row + 1):

        new_entity = {}

        for j in range(1, work_sheet.max_column + 1):

            key_name = work_sheet.cell(row=1, column=j).value
            cell_value = work_sheet.cell(row=i, column=j).value

            if work_sheet.title == 'animals' and cell_value is None:
                continue
            else:
                new_entity[key_name] = cell_value

        try:
            entities.append(new_entity)
        except UnboundLocalError:
            print(f"The sheet '{sheet_name}' is empty, add the data first")
            return []

    return entities


def get_users_with_associated_cages(users_list: list) -> list:
    
    result = []

    for user in users_list:

        user_id = user['id']

        cages = user['responsible_cages']

        if cages == '' or cages is None:
            cages = []
        
        else:
            cages = cages.split(";")

            try:
                cages.remove('')
            except ValueError:
                print("ONE ELEMENT")

            cages = list(map(int, cages))
        
        entry = {
            "user_id": user_id,
            "cages": cages
        }

        result.append(entry)
    
    return result


def get_sections_with_associated_cages(sections_list: list):

    result = []

    for section in sections_list:

        section_id = section['id']

        cages = section['cages']

        print(f"cages: {cages}")

        if cages == '' or cages is None:
            cages = []
        
        else:
            cages = cages.split(";")

            cages.remove('')

            cages = list(map(int, cages))

        entry = {
            "section_id": section_id,
            "cages": cages
        }

        result.append(entry)

    return result


def get_cages_with_associated_animals(cages_list: list):
    
    result = []

    for cage in cages_list:

        cage_id = cage['id']

        animals = cage['animals']

        if animals == '' or animals is None:

            animals = []
        
        else:
            try:
                animals = animals.split(";")
            except AttributeError:
                animals = list(map(str, animals))

            try:
                animals.remove('')
            except ValueError:
                print("Please end the list of animal IDs with ';'")

            animals = list(map(int, animals))
        
        entry = {
            "cage_id": cage_id,
            "animals": animals
        }

        result.append(entry)
    
    return result
    

def save_users_to_file(sheet_name: str, work_book: Workbook, data: list[User], file_path: str):

    # if sheet_name not in work_book.sheetnames:
    #     print(f"The sheet '{sheet_name}' doesn't exist")
    #     return []

    work_sheet = work_book[sheet_name]

    formatted_data = [
        ['id', 'name', 'role', 'responsible_cages', 'shift_is_active']
    ]

    for user in data:
        new_row = [user.id, user.name, user.role, user.responsible_cages, user.shift_is_active]

        formatted_data.append(new_row)

    
    print(f"FORMATTED DATA: {formatted_data}")


    work_sheet.delete_rows(1, work_sheet.max_row)

    for row_index, row_data in enumerate(formatted_data, start=1):
        print(row_data)

        for index, item in enumerate(row_data, start=1):

            if item == []:
                work_sheet.cell(row= row_index, column= index, value=' ')
            else:
                work_sheet.cell(row= row_index, column= index, value=item)
    
    work_book.save(file_path)
    








        




